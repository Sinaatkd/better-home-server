import static_variables

from datetime import datetime

from django.shortcuts import render
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse, reverse_lazy
from django.contrib.auth import get_user_model
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView

from openpyxl import Workbook

from .forms import CreateUserForm, UpdateUserForm, ChangeUserPasswordForm

from estate_module.models import Estate
from user_contact_module.models import UserContact
from user_income_module.models import UserIncome
from user_income_module.utils import calculate_income_general_stats

User = get_user_model()


class AllUsersListView(ListView):
    model = User
    paginate_by = 30
    template_name = 'user/user_list.html'

    def get_queryset(self):
        search = self.request.GET.get('search', '')
        object_list = self.model.objects.all().order_by('-id')
        if search:
            object_list = self.model.objects.search_user_by_username_or_phone_number(search).order_by('-id')
        return object_list

class CreateUserView(CreateView):
    model = User
    form_class = CreateUserForm
    template_name = 'user/user_form.html'

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.set_password(obj.password)
        obj.save()
        return HttpResponseRedirect(reverse('users-list'))


class ConsultantUsersListView(ListView):
    model = User
    paginate_by = 30
    template_name = 'user/user_list.html'

    def get_queryset(self):
        search = self.request.GET.get('search', '')
        object_list = self.model.objects.filter(is_consultant=True).order_by('-id')
        if search:
            object_list = self.model.objects.search_user_by_username_or_phone_number(search).filter(is_consultant=True).order_by('-id')
        return object_list

class UserUpdateView(UpdateView):
    model = User
    form_class = UpdateUserForm
    template_name = 'user/user_form.html'
    success_url = reverse_lazy('users-list')

class UserDetailView(DetailView):
    model = User
    template_name = 'user/user_detail.html'

    def get_context_data(self, **kwargs):
        context = kwargs
        user_id = self.kwargs['pk']
        
        user_incomes = UserIncome.objects.filter(user_id=user_id)
        context['user_incomes'] = user_incomes
        
        sum_of_incomes, this_month_income, progress = calculate_income_general_stats(user_incomes)
        
        context['user_incomes_sum_of_incomes'] = sum_of_incomes
        context['user_incomes_this_month_income'] = this_month_income
        context['user_incomes_progress'] = progress

        fav_estates = Estate.objects.filter(fav_of_users__in=[user_id])
        context['fav_estates'] = fav_estates

        estates = Estate.objects.filter(consultant_id=user_id)
        context['estates'] = estates
        return context
    

class ChangeUserPasswordUpdateView(UpdateView):
    model = User
    form_class = ChangeUserPasswordForm
    template_name = 'user/user_form.html'

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.set_password(obj.password)
        obj.save()
        return HttpResponseRedirect(reverse('users-list'))


class DeleteUserView(DeleteView):
    model = User
    template_name = 'user/user_confirm_delete.html'
    success_url = reverse_lazy('users-list')

    def get_context_data(self, **kwargs):
        consultants = User.objects.filter(is_active=True, is_consultant=True)
        kwargs['consultants']= consultants
        return kwargs

def export_user_to_xlsx(request):
    """
    Downloads all Users as Excel file with a single worksheet
    """
    users_queryset = User.objects.all()
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    date_now = datetime.now()
    response['Content-Disposition'] = f'attachment; filename=users-{date_now.date()}.xlsx'
    workbook = Workbook()
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Users'

    # Define the titles for columns
    columns = static_variables.EXPORT_USER_XLSX_CELL_FIELDS

    row_num = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all users
    for user in users_queryset:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            user.pk,
            user.is_superuser,
            user.username,
            user.first_name,
            user.last_name,
            user.email,
            user.is_staff,
            user.is_active,
            user.date_joined.date(),
            user.country_code,
            user.phone_number,
            user.is_phone_number_verified,
            user.is_consultant,
            user.ad_monthly_quota,
            user.ladder_monthly_quota,
            user.special_ad_monthly_quota,
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def move_consultant_estates_to_other_consultant(request, pk):
    if request.POST:
        new_consultant_id = request.POST.get('consultant')
        estates = Estate.objects.filter(consultant=pk)
        for estate in estates:
            estate.consultant_id = new_consultant_id
            estate.save()
        
        consultant_contacts = UserContact.objects.filter(consultant_id=pk)
        for consultant_contact in consultant_contacts:
            consultant_contact.consultant_id = new_consultant_id
            consultant_contact.save()
            
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))